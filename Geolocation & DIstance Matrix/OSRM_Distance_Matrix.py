import pandas as pd
import requests
from openpyxl import load_workbook
from tqdm import tqdm

# OSRM Server URL (ensure it's running)
OSRM_URL = "http://localhost:5000/route/v1/driving"

# Load Excel File
excel_file = "distance_matrix.xlsx"  # Update with your actual file name
wb = load_workbook(excel_file)
ws = wb.active

# Extract row and column headers (coordinates)
row_coords = [ws[f"C{row}"].value for row in range(4, ws.max_row + 1)]  # C4:C{last row}
col_coords = [ws.cell(row=3, column=col).value for col in range(4, ws.max_column + 1)]  # D3:{last column}

# Function to get distance from OSRM
def get_osrm_distance(coord1, coord2):
    url = f"{OSRM_URL}/{coord1};{coord2}?overview=false"
    try:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            distance_meters = data["routes"][0]["distance"]  # Distance in meters
            distance_km = distance_meters / 1000  # Convert meters to kilometers
            return round(distance_km, 2)  # Round to 2 decimal places
        else:
            print(f"Error {response.status_code}: {response.text}")
            return None
    except Exception as e:
        print(f"Request failed: {e}")
        return None

# Loop through the matrix and fill distances
for row_idx, start_coord in tqdm(enumerate(row_coords, start=4), total=len(row_coords)):
    for col_idx, end_coord in enumerate(col_coords, start=4):
        if start_coord and end_coord:  # Ensure coordinates exist
            distance = get_osrm_distance(start_coord, end_coord)
            if distance is not None:
                ws.cell(row=row_idx, column=col_idx, value=distance)

# Save updated Excel file
wb.save("updated_" + excel_file)
print("Distance matrix computation complete. Saved as updated_" + excel_file)
